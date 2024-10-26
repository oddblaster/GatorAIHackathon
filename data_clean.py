import xlwt
import xlrd
from xlutils.copy import copy
from openpyxl import Workbook, load_workbook
import string





# Load the existing workbook
wb = load_workbook('Florida_DemographicsByCity_sample.xlsx')  # Replace 'your_file.xlsx' with your actual file name

# Load the existing workbook
# Select a specific sheet (by name or index)
sheet = wb.active  # This selects the active sheet
# sheet = wb['SheetName']  # Uncomment this to select a sheet by name



new_wb = Workbook()
new_sheet = new_wb.active
for i in range(1, sheet.max_row + 1):
    new_sheet.cell(row = i, column= 1, value=sheet.cell(row=i + 4,column=1).value)
    new_sheet.cell(row = i, column=2, value=sheet.cell(row=i + 4,column=6).value)  
new_wb.save('formatted_file.xlsx')
